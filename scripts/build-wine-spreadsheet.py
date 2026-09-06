#!/usr/bin/env python3
"""Build CruAtlas World Wine Regions workbook."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins

ROOT = Path("/workspace")
OUT = ROOT / "artifacts" / "CruAtlas-World-Wine-Regions.xlsx"

WINE = "6B1D2A"
WINE_DK = "3F1018"
CREAM = "FBF6EE"
CREAM_DK = "F1E6D4"
INK = "1C1917"
SAGE = "5C6B52"
SAGE_LT = "E7EDE3"
SLATE = "3D4A3A"
GOLD = "C4A35A"
WHITE = "FFFFFF"
MUTED = "6B6560"
ROW_ALT = "F7F1E3"
YES_GREEN = "E4F0D8"
NO_GREY = "F0EBE4"

thin = Border(
    left=Side(style="thin", color="E4D8C8"),
    right=Side(style="thin", color="E4D8C8"),
    top=Side(style="thin", color="E4D8C8"),
    bottom=Side(style="thin", color="E4D8C8"),
)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(size=11, bold=False, color=INK, italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_atlas():
    text = (ROOT / "src/data/countries.ts").read_text()
    cpat = re.compile(r'\n  \{\n    id: "([^"]+)",\n    name: "([^"]+)",\n    flag: "([^"]+)"')
    countries_meta = [(m.start(), m.group(1), m.group(2), m.group(3)) for m in cpat.finditer(text)]

    rpat = re.compile(
        r'id:\s*"([^"]+)",\n\s+name:\s*"([^"]+)",\n\s+lat:\s*([-\d.]+),\n\s+lng:\s*([-\d.]+),\n'
        r'\s+climate:\s*"(.*?)",\n'
        r'\s+soil:\s*"(.*?)",\n'
        r'\s+water:\s*"(.*?)",\n'
        r'\s+vines:\s*\[(.*?)\],\n'
        r'(.*?)\n\s+desc:\s*"(.*?)"',
        re.S,
    )
    rows = []
    for m in rpat.finditer(text):
        pos = m.start()
        cid = cname = flag = ""
        for start, i, n, f in countries_meta:
            if start < pos:
                cid, cname, flag = i, n, f
            else:
                break
        vines = [v.strip() for v in re.findall(r'"([^"]+)"', m.group(8))]
        wine_block = m.group(9)
        wines = re.findall(r'name:\s*"([^"]+)"', wine_block)
        rows.append(
            {
                "country_id": cid,
                "country": cname,
                "flag": flag,
                "id": m.group(1),
                "name": m.group(2),
                "lat": float(m.group(3)),
                "lng": float(m.group(4)),
                "climate": m.group(5).replace('\\"', '"'),
                "soil": m.group(6).replace('\\"', '"'),
                "water": m.group(7).replace('\\"', '"'),
                "vines": vines,
                "wines": wines,
                "desc": m.group(10).replace('\\"', '"'),
            }
        )
    return countries_meta, rows


def load_geojson(path: Path):
    data = json.loads(path.read_text())
    out = []
    for ft in data.get("features", []):
        p = ft.get("properties") or {}
        grapes = p.get("grapes") or []
        if isinstance(grapes, str):
            grapes = [g.strip() for g in grapes.split(",") if g.strip()]
        grapes = [clean_grape(g) for g in grapes if g]
        out.append(
            {
                "id": p.get("id") or "",
                "name": p.get("name") or p.get("full_name") or "",
                "full_name": p.get("full_name") or p.get("name") or "",
                "region": p.get("region") or "",
                "grapes": grapes,
                "key_appellations": p.get("key_appellations") or [],
                "desc": p.get("desc") or "",
                "level": p.get("level") or "appellation",
            }
        )
    return out


def clean_grape(g: str) -> str:
    if not g:
        return g
    g = g.replace("Muscat À Petits Grainslancs", "Muscat à Petits Grains Blancs")
    g = g.replace("Muscat À Petits Grains Roses", "Muscat à Petits Grains Roses")
    g = g.replace("Pinotlanc", "Pinot Blanc")
    g = g.replace("Pinot Grislanc", "Pinot Gris")
    g = g.replace("Pinot Noirlanc", "Pinot Noir")
    g = g.strip()
    g = re.sub(r"\s+", " ", g)
    return g


# ---------------------------------------------------------------------------
# World catalog (regions not already covered by FR/IT geojson + atlas macros)
# Compact TSV: country | continent | macro | region | class | grapes | style | climate
# ---------------------------------------------------------------------------

WORLD_TSV = r"""
Austria	Europe	Niederösterreich	Wachau	DAC	Grüner Veltliner, Riesling	Dry, mineral whites of world class	Cool continental, Danube terraces
Austria	Europe	Niederösterreich	Kamptal	DAC	Grüner Veltliner, Riesling	Structured dry whites, Heiligenstein Riesling	Cool continental
Austria	Europe	Niederösterreich	Kremstal	DAC	Grüner Veltliner, Riesling	Ripe yet precise whites	Cool continental
Austria	Europe	Niederösterreich	Traisental	DAC	Grüner Veltliner, Riesling	Lean, spicy whites	Cool continental
Austria	Europe	Niederösterreich	Wagram	DAC	Grüner Veltliner, Roter Veltliner	Loess-driven, full whites	Cool continental
Austria	Europe	Niederösterreich	Weinviertel	DAC	Grüner Veltliner	Peppery classic Grüner	Cool continental
Austria	Europe	Niederösterreich	Carnuntum	DAC	Zweigelt, Blaufränkisch, Merlot	Ripe, spicy reds east of Vienna	Pannonian
Austria	Europe	Niederösterreich	Thermenregion	DAC	Zierfandler, Rotgipfler, Pinot Noir, St. Laurent	Historic whites and elegant reds	Pannonian
Austria	Europe	Burgenland	Neusiedlersee	DAC	Zweigelt, Welschriesling	Red and sweet botrytis wines	Warm Pannonian, lake
Austria	Europe	Burgenland	Leithaberg	DAC	Blaufränkisch, Grüner Veltliner, Pinot Blanc	Limestone reds and whites	Pannonian
Austria	Europe	Burgenland	Mittelburgenland	DAC	Blaufränkisch	Austria's red-wine heartland	Pannonian
Austria	Europe	Burgenland	Eisenberg	DAC	Blaufränkisch	Iron-rich, mineral reds	Pannonian
Austria	Europe	Steiermark	Südsteiermark	DAC	Sauvignon Blanc, Gelber Muskateller, Morillon	Racy, aromatic whites	Cool alpine-influenced
Austria	Europe	Steiermark	Vulkanland Steiermark	DAC	Sauvignon Blanc, Traminer	Volcanic whites	Cool
Austria	Europe	Steiermark	Weststeiermark	DAC	Blauer Wildbacher (Schilcher)	Rosé specialty	Cool
Austria	Europe	Wien	Wiener Gemischter Satz	DAC	Field blend whites	Vienna's mixed-planted whites	Cool urban vineyards
Hungary	Europe	Tokaj	Tokaj	PDO	Furmint, Hárslevelű, Sárgamuskotály	Aszú dessert wines and dry Furmint	Continental, botrytis
Hungary	Europe	Northern Hungary	Eger	PDO	Kékfrankos, Kadarka, Cabernet, Merlot	Egri Bikavér (Bull's Blood)	Cool continental
Hungary	Europe	Southern Transdanubia	Villány	PDO	Cabernet Franc, Merlot, Portugieser	Hungary's Bordeaux-style reds	Warm continental
Hungary	Europe	Southern Transdanubia	Szekszárd	PDO	Kékfrankos, Kadarka, Merlot	Elegant Bikavér, Kadarka	Warm continental
Hungary	Europe	Balaton	Badacsony	PDO	Olaszrizling, Kéknyelű	Volcanic whites on lake basalt	Lake-moderated
Hungary	Europe	Balaton	Balatonfüred-Csopak	PDO	Olaszrizling	Lakeside whites	Lake-moderated
Hungary	Europe	Northwest	Somló	PDO	Juhfark, Furmint, Hárslevelű	Tiny volcanic hill, long-lived whites	Continental
Hungary	Europe	Northern Transdanubia	Neszmély	PDO	Irsai Olivér, Sauvignon Blanc	Aromatic whites	Cool
Greece	Europe	Cyclades	Santorini	PDO	Assyrtiko, Aidani, Athiri	Assyrtiko on volcanic pumice; Vinsanto	Hot, windy, arid
Greece	Europe	Peloponnese	Nemea	PDO	Agiorgitiko	Velvet reds, 'Blood of Heracles'	Mediterranean, altitude
Greece	Europe	Macedonia	Naoussa	PDO	Xinomavro	Nebbiolo-like structured reds	Continental
Greece	Europe	Macedonia	Amyndeon	PDO	Xinomavro	High-altitude, lighter Xinomavro, sparkling	Cool, lakes
Greece	Europe	Macedonia	Drama	PGI	Sauvignon Blanc, Cabernet, Agiorgitiko	Modern estate wines	Continental
Greece	Europe	Macedonia	Goumenissa	PDO	Xinomavro, Negoska	Softer northern reds	Continental
Greece	Europe	Peloponnese	Mantinia	PDO	Moschofilero	Floral, pink-skinned white/rosé	High plateau
Greece	Europe	Peloponnese	Patras	PDO	Roditis, Mavrodaphne, Muscat	Dry whites and sweet Mavrodaphne	Mediterranean
Greece	Europe	Crete	Peza / Archanes / Dafnes / Sitia	PDO	Vidiano, Vilana, Kotsifali, Liatiko, Mandilari	Island revival, indigenous grapes	Mediterranean
Greece	Europe	Epirus	Zitsa	PDO	Debina	Sparkling and still whites	Cool mountain
Greece	Europe	Aegean	Samos	PDO	Muscat Blanc	Sweet Muscat, historic	Mediterranean
Greece	Europe	Aegean	Paros / Rhodes / Lemnos	PDO	Monemvasia, Athiri, Muscat of Alexandria	Island whites and sweet wines	Mediterranean
Switzerland	Europe	Valais	Valais	AOC	Petite Arvine, Fendant, Cornalin, Syrah	Alpine whites and distinctive reds	Dry inner-alpine
Switzerland	Europe	Vaud	Lavaux / Chablais / La Côte	AOC	Chasselas (Fendant)	Terraced UNESCO Chasselas	Lake Geneva
Switzerland	Europe	Geneva	Geneva	AOC	Chasselas, Gamay, Pinot Noir	Fresh, everyday Swiss	Lake-moderated
Switzerland	Europe	Ticino	Ticino	DOC	Merlot	Italian-speaking Merlot country	Southern alpine
Switzerland	Europe	Graubünden	Bündner Herrschaft	AOC	Pinot Noir (Blauburgunder)	Switzerland's finest Pinot	Alpine valley
Switzerland	Europe	Three Lakes	Neuchâtel / Bienne / Vully	AOC	Pinot Noir, Chasselas, Pinot Gris	Pale Pinot and Oeil-de-Perdrix	Lake-moderated
United Kingdom	Europe	England	West Sussex	PDO	Chardonnay, Pinot Noir, Pinot Meunier	Traditional-method sparkling	Cool maritime
United Kingdom	Europe	England	Kent	PDO	Chardonnay, Pinot Noir, Pinot Meunier	Sparkling, rising still wines	Cool maritime
United Kingdom	Europe	England	Hampshire	PDO	Chardonnay, Pinot Noir	South Downs chalk sparkling	Cool maritime
United Kingdom	Europe	England	Essex	Regional	Pinot Noir, Chardonnay, Bacchus	Still Pinot and Bacchus	Cool, drier east
United Kingdom	Europe	England	Surrey	Regional	Chardonnay, Pinot Noir	Sparkling on greensand	Cool maritime
Germany	Europe	Rheinhessen	Rheinhessen	Anbaugebiet	Riesling, Silvaner, Pinot Blanc, Pinot Noir	Germany's largest region; Roter Hang, Wonnegau	Cool continental
Germany	Europe	Pfalz	Pfalz	Anbaugebiet	Riesling, Pinot Noir, Pinot Blanc	Ripe Riesling, Deutsche Weinstraße	Warmer, dry
Germany	Europe	Nahe	Nahe	Anbaugebiet	Riesling	Mineral Riesling, volcanic & slate	Cool
Germany	Europe	Baden	Kaiserstuhl	Bereich	Pinot Noir, Pinot Gris, Pinot Blanc	Germany's warmest; volcanic Pinot	Warm, dry
Germany	Europe	Baden	Baden	Anbaugebiet	Pinot Noir, Müller-Thurgau, Pinot Gris	Pinot country, long growing season	Warm
Germany	Europe	Franken	Franken	Anbaugebiet	Silvaner, Riesling, Müller-Thurgau	Dry Silvaner in Bocksbeutel	Continental
Germany	Europe	Württemberg	Württemberg	Anbaugebiet	Trollinger, Lemberger, Pinot Noir, Riesling	Red-wine specialist	Warm continental
Germany	Europe	Ahr	Ahr	Anbaugebiet	Pinot Noir	Tiny northern Pinot valley	Cool, slate
Germany	Europe	Mittelrhein	Mittelrhein	Anbaugebiet	Riesling	UNESCO gorge Riesling	Cool river
Germany	Europe	Hessische Bergstraße	Hessische Bergstraße	Anbaugebiet	Riesling, Pinot Gris	Tiny, early-ripening	Mild
Germany	Europe	Saale-Unstrut	Saale-Unstrut	Anbaugebiet	Müller-Thurgau, Pinot Blanc, Riesling	Northernmost quality region	Cool continental
Germany	Europe	Sachsen	Sachsen	Anbaugebiet	Müller-Thurgau, Riesling, Pinot Blanc	Elbe terraces around Dresden	Cool continental
Germany	Europe	Mosel	Saar	Bereich	Riesling	Steely, electric Riesling	Cool slate
Germany	Europe	Mosel	Ruwer	Bereich	Riesling	Delicate, floral Riesling	Cool slate
Portugal	Europe	Norte	Vinho Verde	DOC	Alvarinho, Loureiro, Arinto, Trajadura, Vinhão	Light, zippy whites (and red Vinhão)	Cool Atlantic
Portugal	Europe	Norte	Dão	DOC	Touriga Nacional, Encruzado, Jaen, Alfrocheiro	Elegant highland reds and Encruzado	Continental, granite
Portugal	Europe	Centro	Bairrada	DOC	Baga, Bical, Maria Gomes	Tannic Baga, sparkling	Atlantic
Portugal	Europe	Lisboa	Lisboa	DOC	Arinto, Fernão Pires, Touriga Nacional, Castelão	Diverse, great value	Atlantic
Portugal	Europe	Tejo	Tejo	DOC	Fernão Pires, Castelão, Touriga Nacional	Riverside value wines	Warm
Portugal	Europe	Península de Setúbal	Setúbal	DOC	Moscatel de Setúbal, Castelão	Fortified Muscat, Palmela reds	Warm maritime
Portugal	Europe	Algarve	Algarve	DOC	Touriga Nacional, Syrah, Arinto	Southern holiday wines	Hot Mediterranean
Portugal	Europe	Madeira	Madeira	DOC	Sercial, Verdelho, Bual, Malvasia, Tinta Negra	Historic fortified, volcanic	Subtropical
Portugal	Europe	Açores	Pico / Biscoitos / Graciosa	DOC	Verdelho, Arinto dos Açores, Terrantez do Pico	Azores volcanic whites	Atlantic islands
Portugal	Europe	Norte	Távora-Varosa	DOC	Malvasia Fina, Touriga Nacional	Sparkling highland	Cool
Portugal	Europe	Norte	Trás-os-Montes	DOC	Touriga Nacional, Tinta Roriz	Remote mountain reds	Extreme continental
Portugal	Europe	Centro	Beira Interior	DOC	Síria, Fonte Cal, Rufete	Frontier highlands	Continental
Spain	Europe	Castilla y León	Toro	DO	Tinta de Toro (Tempranillo)	Powerful, inky reds	Hot continental
Spain	Europe	Castilla y León	Bierzo	DO	Mencía, Godello	Atlantic-influenced reds and Godello	Cooler, slate
Spain	Europe	Castilla y León	Rueda	DO	Verdejo, Sauvignon Blanc	Spain's everyday aromatic white	Continental
Spain	Europe	Castilla y León	Cigales	DO	Tempranillo, Garnacha	Rosado tradition, rising reds	Continental
Spain	Europe	Castilla y León	Arribes	DO	Juan García, Malvasía	Douro-side indigenous	Hot river canyon
Spain	Europe	Galicia	Ribeiro	DO	Treixadura, Torrontés, Godello, Caiño	Atlantic blends	Wet, mild
Spain	Europe	Galicia	Ribeira Sacra	DO	Mencía, Godello	Heroic terraces on Sil and Miño	Atlantic-continental
Spain	Europe	Galicia	Valdeorras	DO	Godello, Mencía	Godello revival	Inland Galicia
Spain	Europe	Galicia	Monterrei	DO	Godello, Treixadura, Mencía	Warmest Galician DO	Inland, drier
Spain	Europe	Catalunya	Cava	DO	Macabeo, Xarel·lo, Parellada, Chardonnay	Traditional-method sparkling	Mediterranean
Spain	Europe	Catalunya	Penedès	DO	Xarel·lo, Macabeo, Tempranillo, International	Still and sparkling heartland	Mediterranean
Spain	Europe	Catalunya	Montsant	DO	Garnacha, Cariñena	Priorat's neighbour, better value	Warm, slate and clay
Spain	Europe	Catalunya	Empordà	DO	Garnacha, Cariñena	Tramontana-swept, French border	Mediterranean, windy
Spain	Europe	Catalunya	Costers del Segre	DO	Tempranillo, Cabernet, Garnacha	Inland Catalonia, Raimat	Continental-Mediterranean
Spain	Europe	Catalunya	Terra Alta	DO	Garnacha Blanca, Garnacha	High inland whites	Dry, windy
Spain	Europe	Catalunya	Conca de Barberà	DO	Trepat, Macabeo	Trepat rosado, Cava grapes	Mediterranean, altitude
Spain	Europe	Catalunya	Alella	DO	Pansa Blanca (Xarel·lo)	Tiny coastal DO by Barcelona	Mediterranean
Spain	Europe	Navarra	Navarra	DO	Garnacha, Tempranillo, Chardonnay	Rosado classic, modern reds	Atlantic to Mediterranean
Spain	Europe	Aragón	Cariñena	DO	Garnacha, Cariñena, Tempranillo	Old-vine Garnacha	Continental
Spain	Europe	Aragón	Calatayud	DO	Garnacha	High old-vine Garnacha	Continental, altitude
Spain	Europe	Aragón	Campo de Borja	DO	Garnacha	'Empire of Garnacha'	Continental
Spain	Europe	Aragón	Somontano	DO	Cabernet, Chardonnay, Gewürztraminer, Moristel	Pyrenean-foot international	Continental
Spain	Europe	País Vasco	Getariako Txakolina	DO	Hondarrabi Zuri	Spritzy, saline coastal white	Cool Atlantic
Spain	Europe	País Vasco	Bizkaiko Txakolina	DO	Hondarrabi Zuri	Slightly riper Txakoli	Atlantic
Spain	Europe	País Vasco	Arabako Txakolina	DO	Hondarrabi Zuri	Inland, structured Txakoli	Atlantic
Spain	Europe	Andalucía	Jerez-Xérès-Sherry	DO	Palomino, Pedro Ximénez, Moscatel	Fortified: Fino to PX	Hot, albariza chalk
Spain	Europe	Andalucía	Manzanilla-Sanlúcar de Barrameda	DO	Palomino	Coastal biologically aged Fino	Atlantic, humid
Spain	Europe	Andalucía	Montilla-Moriles	DO	Pedro Ximénez	Unfortified and PX sweets	Hot inland
Spain	Europe	Andalucía	Málaga / Sierras de Málaga	DO	Moscatel, Pedro Ximénez, Romé	Sweet and mountain dry	Mediterranean, mountains
Spain	Europe	La Mancha	La Mancha	DO	Airén, Tempranillo (Cencibel), Macabeo	Europe's largest DO, value	Hot continental
Spain	Europe	La Mancha	Valdepeñas	DO	Tempranillo, Airén	Softer La Mancha neighbour	Hot continental
Spain	Europe	Valencia	Utiel-Requena	DO	Bobal	Bobal reds and rosados	Continental plateau
Spain	Europe	Valencia	Valencia	DO	Monastrell, Moscatel, Merseguera	Mediterranean volume and Moscatel	Warm
Spain	Europe	Valencia	Alicante	DO	Monastrell, Moscatel	Fondillón, Mediterranean reds	Hot, dry
Spain	Europe	Murcia	Jumilla	DO	Monastrell	Ripe, sunny Monastrell	Hot, arid
Spain	Europe	Murcia	Yecla	DO	Monastrell	High-altitude Monastrell	Hot, arid
Spain	Europe	Murcia	Bullas	DO	Monastrell	Cooler Murcian highlands	Continental
Spain	Europe	Madrid	Vinos de Madrid	DO	Garnacha, Albillo Real	Sierra de Gredos Garnacha	Continental, granite
Spain	Europe	Extremadura	Ribera del Guadiana	DO	Tempranillo, Macabeo, Pardina	Value interior	Hot
Spain	Europe	Balearics	Binissalem	DO	Manto Negro, Callet, Prensal Blanc	Mallorca indigenous	Mediterranean
Spain	Europe	Balearics	Pla i Llevant	DO	Callet, Fogoneu, Prensal	Eastern Mallorca	Mediterranean
Spain	Europe	Canary Islands	Tacoronte-Acentejo	DO	Listán Negro, Listán Blanco	Tenerife volcanic	Atlantic subtropical
Spain	Europe	Canary Islands	Lanzarote	DO	Malvasía Volcánica	Vines in ash pits (hoyos)	Arid volcanic
Spain	Europe	Canary Islands	La Palma / El Hierro / Gran Canaria / Abona / Valle de la Orotava	DO	Listán Blanco, Negramoll, Vijariego	Island volcanic terroirs	Atlantic
Slovenia	Europe	Primorska	Goriška Brda	PGI/PTP	Rebula, Rebula, Sauvignon, Merlot, Pinot Grigio	Collio continuation, orange-wine culture	Sub-Mediterranean
Slovenia	Europe	Primorska	Vipava Valley	PTP	Zelen, Pinela, Malvazija	Indigenous whites, windy valley	Sub-Mediterranean
Slovenia	Europe	Primorska	Koper	PTP	Refošk (Refosco), Malvazija	Coastal reds	Mediterranean
Slovenia	Europe	Podravje	Štajerska	PTP	Furmint (Šipon), Riesling, Sauvignon, Pinot Blanc	Cool aromatic whites	Continental
Slovenia	Europe	Posavje	Bizeljsko / Dolenjska	PTP	Modra Frankinja, Yellow Muscat	Cviček blend, sparkling	Continental
Croatia	Europe	Dalmatia	Dingač / Postup (Pelješac)	PDO	Plavac Mali	Croatia's most powerful reds	Hot Mediterranean
Croatia	Europe	Dalmatia	Hvar / Korčula / Brač	PDO	Pošip, Bogdanuša, Plavac Mali	Island whites and reds	Mediterranean
Croatia	Europe	Istria	Istria	PDO	Malvazija Istarska, Teran	Italian-border whites and Teran	Sub-Mediterranean
Croatia	Europe	Slavonia	Kutjevo / Baranja	PDO	Graševina (Welschriesling)	Continental whites	Continental
Croatia	Europe	Croatian Uplands	Plešivica / Zagorje	PDO	Pinot Noir, Chardonnay, Riesling, Škrlet	Sparkling and cool still	Cool continental
Georgia	Europe	Kakheti	Kakheti	PDO	Rkatsiteli, Saperavi, Mtsvane, Kisi	Qvevri amber and Saperavi reds	Continental
Georgia	Europe	Kakheti	Kindzmarauli	PDO	Saperavi	Naturally semi-sweet red	Continental
Georgia	Europe	Kakheti	Mukuzani	PDO	Saperavi	Dry, structured Saperavi	Continental
Georgia	Europe	Kakheti	Tsinandali	PDO	Rkatsiteli, Mtsvane	Classic Kakhetian white	Continental
Georgia	Europe	Kartli	Kartli	PDO	Chinuri, Goruli Mtsvane, Pinot Noir	Sparkling and qvevri	Continental
Georgia	Europe	Imereti	Imereti	PDO	Tsolikouri, Tsitska, Krakhuna	Western, fresher whites	Humid, milder
Georgia	Europe	Racha-Lechkhumi	Khvanchkara / Tvishi	PDO	Aleksandrouli, Mujuretuli, Tsolikouri	Semi-sweet reds, Tvishi white	Mountain
Romania	Europe	Moldova (RO)	Dealu Mare	DOC	Fetească Neagră, Merlot, Cabernet, Tămâioasă	Romania's premium red hills	Continental
Romania	Europe	Transylvania	Târnave	DOC	Fetească Albă, Fetească Regală, Riesling	Cool whites, sparkling	Cool continental
Romania	Europe	Dobrogea	Murfatlar	DOC	Chardonnay, Pinot Gris, Muscat, Cabernet	Black Sea, sweet and dry	Maritime-continental
Romania	Europe	Oltenia	Drăgășani	DOC	Crâmpoșie, Fetească Neagră, Syrah	Revival region on Olt	Continental
Romania	Europe	Moldova (RO)	Cotnari	DOC	Grasă de Cotnari, Tămâioasă	Historic sweet whites	Continental
Bulgaria	Europe	Thracian Lowlands	Thracian Lowlands	PGI	Mavrud, Rubin, Merlot, Cabernet, Syrah	Southern reds, Mavrud	Warm continental
Bulgaria	Europe	Danube Plain	Danube Plain	PGI	Cabernet, Merlot, Dimyat, Muscat	Northern value wines	Continental
Bulgaria	Europe	Black Sea	Black Sea	PGI	Dimyat, Muscat, Chardonnay	Coastal whites	Maritime
Bulgaria	Europe	Struma Valley	Struma Valley	PGI	Melnik (Broad Leaf), Shiroka Melnishka	SW reds near Greece	Warm, windy
Moldova	Europe	Codru	Codru	IG	Rara Neagră, Fetească, Cabernet, Merlot	Forest-belt reds and whites	Continental
Moldova	Europe	Ștefan Vodă	Purcari / Ștefan Vodă	IG	Rara Neagră, Cabernet, Merlot	Flagship eastern estates	Continental
Moldova	Europe	Valul lui Traian	Valul lui Traian	IG	Cabernet, Merlot, Rara Neagră	Southern, fuller	Continental
Czech Republic	Europe	Moravia	Mikulov	VOC	Riesling, Grüner Veltliner, Pinot Blanc, St. Laurent, Pálava	Moravian whites and ice wine	Cool continental
Czech Republic	Europe	Moravia	Znojmo	VOC	Sauvignon Blanc, Riesling, Grüner Veltliner	Aromatic northern Moravia	Cool
Czech Republic	Europe	Moravia	Velké Pavlovice	VOC	St. Laurent, Blaufränkisch, Riesling	Redder Moravian corner	Cool
Czech Republic	Europe	Moravia	Slovácko	VOC	Riesling, Pinot Gris, Frankovka	SE Moravia	Cool
Slovakia	Europe	Small Carpathians	Malokarpatská	TOC	Grüner Veltliner, Riesling, Blaufränkisch, St. Laurent	Little Carpathian whites and reds	Cool continental
Slovakia	Europe	South Slovakia	Južnoslovenská	TOC	Cabernet, Riesling, Dunaj	Warmer Danube south	Continental
Slovakia	Europe	Tokaj	Tokajská	TOC	Furmint, Lipovina, Muscat	Slovak share of Tokaj	Continental
Luxembourg	Europe	Moselle	Moselle Luxembourgeoise	AOP	Riesling, Pinot Gris, Pinot Blanc, Auxerrois, Gewürztraminer	River terraces, Crémant	Cool
Belgium	Europe	Flanders / Wallonia	Hageland / Côtes de Sambre et Meuse	AOP	Chardonnay, Pinot Noir, Müller-Thurgau	Cool sparkling and still	Cool maritime
Netherlands	Europe	Limburg / Gelderland	Maasvallei / Achterhoek	PDO	Auxerrois, Pinot Gris, Rondo, Johanniter	Nascent cool-climate	Cool maritime
North Macedonia	Europe	Povardarie	Tikveš	PGI	Vranec, Smederevka, Tamjanika	Vranec heartland	Hot continental
Serbia	Europe	Central Serbia	Župa / Šumadija / Fruška Gora	PGI	Prokupac, Tamjanika, Riesling, Frankovka	Revival of Prokupac, Fruška Gora whites	Continental
Montenegro	Europe	Lake Skadar	Crmnica / Lake Skadar	PGI	Vranac, Krstač	Vranac, scenic lake vineyards	Mediterranean-continental
Bosnia and Herzegovina	Europe	Herzegovina	Mostar / Čitluk	PGI	Žilavka, Blatina	Žilavka whites, Blatina reds	Mediterranean inland
Albania	Europe	Western Lowlands	Durrës / Berat / Përmet	Regional	Shesh i Zi, Shesh i Bardhë, Kallmet	Indigenous Adriatic wines	Mediterranean
Kosovo	Europe	Rahovec	Rahovec / Orahovac	Regional	Vranac, Prokupac, Smellerova	Interior Balkans	Continental
Ukraine	Europe	Black Sea	Odesa / Mykolaiv / Kherson	PGI	Odesa Black, Rkatsiteli, Cabernet, Merlot, Chardonnay	Black Sea steppe, sparkling tradition	Continental, maritime
Ukraine	Europe	Transcarpathia	Zakarpattia	PGI	Furmint, Riesling, Cabernet	Carpathian foothills	Cool continental
Ukraine	Europe	Crimea	Crimea (historic)	Historic	Muscat, Cabernet, Aligoté	Historic Massandra, sparkling	Maritime
Armenia	Europe	Vayots Dzor	Vayots Dzor / Areni	GI	Areni, Voskehat	High-altitude Areni, ancient caves	Continental, high
Armenia	Europe	Ararat	Ararat Valley	GI	Areni, Kangun, Tozot	Ararat-foot vineyards	Continental
Turkey	Europe	Aegean	Denizli / İzmir / Cappadocia / Thrace	Regional	Öküzgözü, Boğazkere, Kalecik Karası, Narince, Emir	Indigenous revival, also international	Mediterranean to continental
Cyprus	Europe	Limassol	Commandaria	PDO	Xynisteri, Mavro	World's oldest named wine, sun-dried	Hot Mediterranean
Cyprus	Europe	Limassol / Paphos	Krasochoria / Laona-Akamas	PDO	Xynisteri, Maratheftiko, Yiánnoudi	Mountain and coastal dry wines	Mediterranean, altitude
Lebanon	Asia	Bekaa Valley	Bekaa Valley	GI	Cinsault, Cabernet, Syrah, Merlot, Viognier, Obaideh	Historic high-valley reds (Château Musar)	High-altitude Mediterranean
Lebanon	Asia	Mount Lebanon	Batroun / Jezzine	GI	Cinsault, Syrah, Viognier	Coastal mountain revival	Mediterranean, altitude
Israel	Asia	Galilee	Upper Galilee	GI	Cabernet, Merlot, Syrah, Chardonnay	High Galilee, flagship quality	Mediterranean, altitude
Israel	Asia	Golan	Golan Heights	GI	Cabernet, Merlot, Sauvignon Blanc, Riesling	Volcanic plateau	Cool high
Israel	Asia	Judean Hills	Judean Hills	GI	Cabernet, Syrah, Chenin, Mediterranean blends	Hillside revival	Mediterranean
Israel	Asia	Coast	Shomron / Samson / Negev	GI	Cabernet, Carignan, Argaman	Coastal and desert drip-irrigated	Mediterranean to arid
United States	North America	California	Oakville	AVA	Cabernet Sauvignon, Merlot	Napa Cabernet benchmark	Warm, dry summers
United States	North America	California	Rutherford	AVA	Cabernet Sauvignon	Dusty Rutherford Cabernet	Warm
United States	North America	California	Stags Leap District	AVA	Cabernet Sauvignon	Soft tannin, polished Cabernet	Warm
United States	North America	California	Howell Mountain	AVA	Cabernet Sauvignon	Mountain intensity	Warm days, elevation
United States	North America	California	Mount Veeder	AVA	Cabernet Sauvignon, Chardonnay	Mayacamas mountain	Cooler, elevation
United States	North America	California	Spring Mountain District	AVA	Cabernet Sauvignon	West Mayacamas	Elevation
United States	North America	California	Atlas Peak	AVA	Cabernet Sauvignon	East hills, volcanic	Elevation
United States	North America	California	Calistoga	AVA	Cabernet Sauvignon, Zinfandel	Warmest Napa AVA	Warm
United States	North America	California	Coombsville	AVA	Cabernet Sauvignon, Syrah	Cooler southern Napa	Cooler
United States	North America	California	Los Carneros	AVA	Pinot Noir, Chardonnay, Merlot	Cool, sparkling and Pinot	Cool, windy
United States	North America	California	Russian River Valley	AVA	Pinot Noir, Chardonnay	Fog-cooled, Goldridge soils	Cool
United States	North America	California	Sonoma Coast	AVA	Pinot Noir, Chardonnay	Extreme coastal	Very cool
United States	North America	California	Dry Creek Valley	AVA	Zinfandel, Sauvignon Blanc, Cabernet	Old-vine Zinfandel	Warm days, cool nights
United States	North America	California	Alexander Valley	AVA	Cabernet Sauvignon	Sonoma Cabernet	Warm
United States	North America	California	Anderson Valley	AVA	Pinot Noir, Chardonnay, Gewürztraminer, Riesling	Mendocino fog pocket, sparkling	Cool
United States	North America	California	Mendocino	AVA	Zinfandel, Pinot Noir, Chardonnay, Petite Sirah	Diverse, organic pioneer	Varied
United States	North America	California	Paso Robles	AVA	Cabernet, Syrah, Rhône varieties, Zinfandel	Central Coast powerhouse	Warm, diurnal
United States	North America	California	Santa Lucia Highlands	AVA	Pinot Noir, Chardonnay, Syrah	Windy benches above Salinas	Cool
United States	North America	California	Sta. Rita Hills	AVA	Pinot Noir, Chardonnay	Transverse range, limestone	Cool, windy
United States	North America	California	Santa Maria Valley	AVA	Pinot Noir, Chardonnay	Long, cool season	Cool
United States	North America	California	Santa Cruz Mountains	AVA	Cabernet, Pinot Noir, Chardonnay	Ridge Monte Bello country	Cool mountain
United States	North America	California	Lodi	AVA	Zinfandel, old-vine mixed	Old-vine Zinfandel, value	Warm
United States	North America	California	Sierra Foothills	AVA	Zinfandel, Barbera, Rhône	Gold Country old vines	Warm, elevation
United States	North America	California	Livermore Valley	AVA	Cabernet, Chardonnay, Sauvignon Blanc	Historic East Bay	Warm
United States	North America	California	Temecula Valley	AVA	Rhône, Mediterranean, sparkling	SoCal AVA	Warm
United States	North America	Oregon	Dundee Hills	AVA	Pinot Noir, Chardonnay	Jory volcanic soils	Cool maritime
United States	North America	Oregon	Eola-Amity Hills	AVA	Pinot Noir, Chardonnay	Van Duzer winds	Cool
United States	North America	Oregon	Ribbon Ridge	AVA	Pinot Noir	Tiny, sedimentary	Cool
United States	North America	Oregon	Chehalem Mountains	AVA	Pinot Noir, Chardonnay	Diverse soils	Cool
United States	North America	Oregon	Yamhill-Carlton	AVA	Pinot Noir	Marine sedimentary	Cool
United States	North America	Oregon	Rogue Valley	AVA	Cabernet, Syrah, Pinot, Tempranillo	Southern Oregon warmth	Warm
United States	North America	Oregon	Umpqua Valley	AVA	Pinot Noir, Riesling, Tempranillo, Syrah	Between Willamette and Rogue	Transitional
United States	North America	Washington	Walla Walla Valley	AVA	Cabernet, Syrah, Merlot	Flagship WA reds	Continental desert
United States	North America	Washington	Red Mountain	AVA	Cabernet Sauvignon, Syrah	Tiny, intense, high heat	Hot, dry
United States	North America	Washington	Horse Heaven Hills	AVA	Cabernet, Merlot, Syrah, Chardonnay	Wind-swept benches	Hot, dry
United States	North America	Washington	Yakima Valley	AVA	Riesling, Cabernet, Syrah, Chardonnay	Oldest WA AVA	Continental
United States	North America	New York	Finger Lakes	AVA	Riesling, Cabernet Franc, Gewürztraminer, Pinot Noir	Lake-moderated Riesling	Cool continental
United States	North America	New York	North Fork of Long Island	AVA	Merlot, Cabernet Franc, Chardonnay, Sauvignon Blanc	Maritime Bordeaux-like	Cool maritime
United States	North America	New York	Hudson River Region	AVA	Hybrid and vinifera, sparkling	Historic, cool	Cool
United States	North America	Virginia	Monticello / Shenandoah / Middleburg	AVA	Viognier, Cabernet Franc, Petit Verdot, Chardonnay, Petit Manseng	Humid, improving reds and Viognier	Humid subtropical
United States	North America	Texas	Texas High Plains / Hill Country	AVA	Tempranillo, Mourvèdre, Viognier, Cabernet	Hot, high, Mediterranean varieties	Hot, dry, altitude
United States	North America	Michigan	Old Mission Peninsula / Leelanau	AVA	Riesling, Pinot Noir, Chardonnay, Cabernet Franc	Lake Michigan snow belt	Cool
United States	North America	New Mexico	Middle Rio Grande / Mesilla	AVA	Sparkling, Riesling, Italian varieties	High desert, Gruet sparkling	High desert
Canada	North America	Ontario	Niagara Peninsula	VQA	Riesling, Chardonnay, Cabernet Franc, Pinot Noir, Vidal	Icewine capital, table wines rising	Cool continental
Canada	North America	Ontario	Prince Edward County	VQA	Pinot Noir, Chardonnay	Limestone island, cool Pinot	Cool
Canada	North America	British Columbia	Okanagan Valley	VQA	Pinot Noir, Riesling, Chardonnay, Syrah, Cabernet	Desert lakeshore, N-S diversity	Continental desert
Canada	North America	British Columbia	Similkameen	GI	Bordeaux reds, Riesling	Windy organic valley	Dry
Canada	North America	Nova Scotia	Annapolis Valley / Tidal Bay	GI	L'Acadie Blanc, Vidal, hybrids	Tidal Bay whites, sparkling	Cool maritime
Mexico	North America	Baja California	Valle de Guadalupe	GI	Cabernet, Tempranillo, Chenin, Grenache, Nebbiolo	Mexico's fine-wine heart	Warm, dry, maritime nights
Mexico	North America	Coahuila	Parras Valley	GI	Cabernet, Shiraz	Oldest American winery tradition	High desert
Argentina	South America	Mendoza	Luján de Cuyo	GI	Malbec, Cabernet Sauvignon	Historic Malbec, first GI	High desert
Argentina	South America	Mendoza	Valle de Uco	GI	Malbec, Cabernet Franc, Chardonnay, Pinot Noir	High Andean, most exciting	High, cool desert
Argentina	South America	Mendoza	Maipú	GI	Malbec, Bonarda, Torrontés, olive country	Historic east Mendoza	Desert
Argentina	South America	Mendoza	Tupungato	GI	Malbec, Chardonnay	Uco, Gualtallary altitude	Very high
Argentina	South America	Patagonia	Neuquén	GI	Malbec, Pinot Noir, Merlot	Southern windy desert	Cool, windy
Argentina	South America	Patagonia	Río Negro	GI	Pinot Noir, Malbec, Merlot, Sémillon	Oldest Patagonian vines	Cool
Argentina	South America	Cuyo	San Juan	GI	Syrah, Bonarda, Malbec, Pedro Giménez	Hotter neighbour of Mendoza	Hot desert
Argentina	South America	Northwest	La Rioja (AR)	GI	Torrontés Riojano	Torrontés volume	Hot, high
Chile	South America	Aconcagua	Casablanca Valley	DO	Sauvignon Blanc, Pinot Noir, Chardonnay	Cool coastal whites and Pinot	Cool maritime
Chile	South America	Aconcagua	San Antonio / Leyda	DO	Sauvignon Blanc, Pinot Noir, Chardonnay, Syrah	Extreme coastal	Very cool
Chile	South America	Aconcagua	Aconcagua Valley	DO	Cabernet, Syrah, Carmenère	Warm Andean-foot reds, coastal Syrah	Warm to cool
Chile	South America	Coquimbo	Limarí Valley	DO	Chardonnay, Pinot Noir, Syrah, Sauvignon Blanc	Limestone, camanchaca fog	Cool desert
Chile	South America	Coquimbo	Elqui Valley	DO	Syrah, Sauvignon Blanc, Pedro Ximénez, Pisco grapes	Ultra-arid, high, starry	Desert, high
Chile	South America	Central Valley	Cachapoal	DO	Cabernet, Carmenère	Rapel reds, Peumo Carmenère	Warm
Chile	South America	Central Valley	Colchagua – Apalta	Zona	Carmenère, Cabernet, Syrah	Horseshoe hills, icon reds	Warm
Chile	South America	Central Valley	Curicó	DO	Cabernet, Sauvignon Blanc, Merlot	Volume and value	Warm
Chile	South America	Central Valley	Maule	DO	País, Carignan (VIGNO), Cabernet	Old dry-farmed bush vines	Warm, dry-farmed
Chile	South America	Southern	Itata	DO	País, Cinsault, Muscat, Cinsault	Historic south, old vines	Cooler, rainier
Chile	South America	Southern	Bío Bío / Malleco	DO	Pinot Noir, Chardonnay, Riesling, Gewürztraminer	Cool south	Cool, wet
Uruguay	South America	Canelones	Canelones	GI	Tannat, Albariño, Merlot	Tannat heartland near Montevideo	Humid maritime
Uruguay	South America	Maldonado	Maldonado / Garzón	GI	Tannat, Albariño, Sauvignon Blanc	Atlantic, cooler	Maritime
Uruguay	South America	Colonia	Colonia	GI	Tannat, Cabernet Franc	River Plate	Humid
Brazil	South America	Serra Gaúcha	Vale dos Vinhedos	DO	Merlot, Chardonnay, Tannat, Italian varieties	Sparkling and Merlot, Italian heritage	Humid highland
Brazil	South America	Campanha	Campanha Gaúcha	GI	Tannat, Cabernet, Merlot	Drier south, still reds	Temperate
Brazil	South America	Pinto Bandeira	Pinto Bandeira	DO	Chardonnay, Pinot Noir	Traditional-method sparkling	Highland
Brazil	South America	Vale do São Francisco	Vale do São Francisco	GI	Syrah, Moscato, Chenin	Tropical, two harvests a year	Tropical semi-arid
Peru	South America	Ica	Ica / Moquegua / Arequipa	GI	Italia, Quebranta, Malbec, Tannat	Pisco country, rising still wine	Coastal desert
Bolivia	South America	Tarija	Tarija / Cintis	GI	Tannat, Malbec, Muscat of Alexandria	World's highest commercial vines (~2,800 m)	High desert
Australia	Oceania	South Australia	McLaren Vale	GI	Shiraz, Grenache, Mourvèdre, Cabernet	Mediterranean GSM, coastal	Warm maritime
Australia	Oceania	South Australia	Adelaide Hills	GI	Sauvignon Blanc, Chardonnay, Pinot Noir, Shiraz	Cool Adelaide ranges	Cool
Australia	Oceania	South Australia	Clare Valley	GI	Riesling, Shiraz, Cabernet	Dry Riesling, historic	Warm days, cool nights
Australia	Oceania	South Australia	Eden Valley	GI	Riesling, Shiraz	Higher, cooler Barossa neighbour	Cooler
Australia	Oceania	South Australia	Langhorne Creek	GI	Cabernet, Shiraz	Lake-moderated, blends	Warm
Australia	Oceania	South Australia	Padthaway / Wrattonbully	GI	Cabernet, Shiraz, Chardonnay	Limestone Coast siblings	Cool-warm
Australia	Oceania	Victoria	Mornington Peninsula	GI	Pinot Noir, Chardonnay	Cool maritime, Melbourne's playground	Cool
Australia	Oceania	Victoria	Heathcote	GI	Shiraz	Cambrian soils, spicy Shiraz	Warm, dry
Australia	Oceania	Victoria	Rutherglen	GI	Muscat, Topaque (Muscadelle), Durif	Iconic fortified stickies	Warm
Australia	Oceania	Victoria	Beechworth	GI	Chardonnay, Shiraz, Nebbiolo	NE Victoria, Giaconda country	Cool-warm
Australia	Oceania	Victoria	Macedon Ranges	GI	Pinot Noir, Chardonnay, sparkling	Cool, high	Cool
Australia	Oceania	Victoria	Geelong	GI	Pinot Noir, Chardonnay, Shiraz	Bellarine and Bannockburn	Cool maritime
Australia	Oceania	Victoria	Gippsland	GI	Pinot Noir, Chardonnay	Bassine cool climate	Cool
Australia	Oceania	Tasmania	Tamar Valley	GI	Pinot Noir, Chardonnay, Riesling, sparkling	Northern Tassie	Cool
Australia	Oceania	Tasmania	Coal River / Derwent / East Coast / Pipers River	GI	Pinot Noir, Chardonnay, Riesling, sparkling	Island cool-climate	Cool maritime
Australia	Oceania	Western Australia	Great Southern	GI	Riesling, Cabernet, Shiraz, Chardonnay	Frankland, Porongurup, Mount Barker	Cool-warm maritime
Australia	Oceania	Western Australia	Swan Valley	GI	Verdelho, Chenin, Shiraz, fortified	Hottest Australian GI, historic	Hot
Australia	Oceania	New South Wales	Orange	GI	Chardonnay, Shiraz, Sauvignon Blanc, Pinot	High-altitude inland	Cool high
Australia	Oceania	New South Wales	Mudgee	GI	Cabernet, Shiraz, Chardonnay	West of Hunter	Warm
Australia	Oceania	New South Wales	Canberra District	GI	Riesling, Shiraz, Pinot Noir	Continental, Clonakilla Shiraz Viognier	Cool continental
Australia	Oceania	Queensland	Granite Belt	GI	Shiraz, Cabernet, Verdelho, alternative varieties	High, granite	Warm high
New Zealand	Oceania	North Island	Hawke's Bay	GI	Cabernet, Merlot, Syrah, Chardonnay	Gimblett Gravels Bordeaux and Syrah	Warm maritime
New Zealand	Oceania	North Island	Martinborough / Wairarapa	GI	Pinot Noir, Sauvignon Blanc, Riesling	Early NZ Pinot pioneer	Cool, windy
New Zealand	Oceania	North Island	Gisborne	GI	Chardonnay, Pinot Gris, Gewürztraminer	'Chardonnay capital', first light	Warm
New Zealand	Oceania	North Island	Auckland / Waiheke	GI	Bordeaux blends, Syrah, Chardonnay	Island and west-coast reds	Warm maritime
New Zealand	Oceania	South Island	Nelson	GI	Sauvignon Blanc, Pinot Noir, Chardonnay, aromatic	Sunniest, artisanal	Cool maritime
New Zealand	Oceania	South Island	Canterbury / Waipara	GI	Pinot Noir, Riesling, Sauvignon, Chardonnay	East-coast, dry	Cool, dry
New Zealand	Oceania	South Island	North Canterbury	GI	Pinot Noir, Riesling	Waipara limestone	Cool
New Zealand	Oceania	South Island	Waitaki Valley	GI	Pinot Noir, Riesling, Pinot Gris	Limestone, Otago-North Otago	Cool continental
South Africa	Africa	Western Cape	Paarl	WO	Shiraz, Cabernet, Chenin, Pinotage	Warmer than Stellenbosch	Warm Mediterranean
South Africa	Africa	Western Cape	Franschhoek	WO	Cabernet, Sauvignon Blanc, Semillon, Huguenot history	Valley of the French	Warm
South Africa	Africa	Western Cape	Constantia	WO	Sauvignon Blanc, Muscat (Vin de Constance), Cabernet	Historic cool peninsula	Cool maritime
South Africa	Africa	Western Cape	Durbanville	WO	Sauvignon Blanc, Merlot	Tygerberg hills, fog	Cool
South Africa	Africa	Western Cape	Darling	WO	Sauvignon Blanc, Chenin, Shiraz	Atlantic, old bush vines	Cool-warm
South Africa	Africa	Western Cape	Walker Bay / Hemel-en-Aarde	WO	Pinot Noir, Chardonnay	Coolest Cape, Burgundian	Cool maritime
South Africa	Africa	Western Cape	Elgin	WO	Sauvignon Blanc, Pinot Noir, Chardonnay, Riesling, Syrah	Apple country, cool	Cool
South Africa	Africa	Western Cape	Robertson	WO	Chardonnay, Colombard, Shiraz, fortified	Lime-rich, 'valley of vines and roses'	Warm, dry
South Africa	Africa	Western Cape	Wellington	WO	Chenin, Shiraz, Pinotage	North of Paarl, nurseries	Warm
South Africa	Africa	Western Cape	Tulbagh	WO	Shiraz, Chenin, Colombard, sparkling	Mountain basin	Warm
South Africa	Africa	Western Cape	Elim / Cape Agulhas	WO	Sauvignon Blanc, Semillon, Syrah	Southernmost Africa, windy	Cool, windy
South Africa	Africa	Western Cape	Piekenierskloof	WO	Grenache, Chenin, old vines	High Citrusdal	Warm high
South Africa	Africa	Western Cape	Olifants River	WO	Colombard, Chenin, Pinotage	North, bulk and old vines	Warm
South Africa	Africa	Western Cape	Klein Karoo	WO	Muscat, fortified, Colombard	Semi-arid, Calitzdorp Port-style	Hot, arid
Morocco	Africa	Meknès / Atlas	Guerrouane / Beni M'Tir / Berkane	AOG	Syrah, Cabernet, Grenache, Vin Gris, Carignan	Atlas-foot reds and gris	Mediterranean, altitude
Tunisia	Africa	Cap Bon	Mornag / Coteaux de Tébourba	AOC	Muscat d'Alexandrie, Carignan, Grenache, Syrah	Muscat and Mediterranean reds	Hot Mediterranean
Algeria	Africa	Coteaux	Coteaux de Mascara / Tlemcen / Zaccar	Regional	Carignan, Cinsault, Alicante, Clairette	Historic colonial volume, now small	Mediterranean
China	Asia	Ningxia	Helan Mountain East	GI	Cabernet Sauvignon, Marselan, Merlot, Chardonnay	China's most awarded region, buried vines in winter	Continental desert, high diurnal
China	Asia	Shandong	Penglai / Yantai	GI	Cabernet, Chardonnay, Riesling	Coastal, humid, first modern wineries	Maritime
China	Asia	Xinjiang	Yanqi / Turpan / Ili	GI	Cabernet, Merlot, Riesling, local table grapes	Desert basins, extreme	Arid continental
China	Asia	Hebei	Huailai / Changli	GI	Cabernet, Chardonnay	Near Beijing, mountain and coast	Continental
China	Asia	Yunnan	Deqin / Shangri-La	GI	Cabernet, Pinot Noir, Petit Manseng	Himalayan vineyards 2,000–2,800 m	High, dry
China	Asia	Shanxi	Taigu	GI	Cabernet, Marselan, Italian Riesling	Loess plateau	Continental
Japan	Asia	Yamanashi	Koshu / Yamanashi	GI	Koshu, Muscat Bailey A, Chardonnay, Merlot, Pinot Noir	Koshu white, rain, pergolas	Humid temperate
Japan	Asia	Nagano	Chikumagawa / Kikyogahara	GI	Merlot, Chardonnay, Pinot Noir, Niagara	Central Alps, cooler	Cool highland
Japan	Asia	Hokkaidō	Yoichi / Tokachi / Sorachi	GI	Pinot Noir, Kerner, Müller-Thurgau, Zweigelt	Cool island, rising Pinot	Cool
Japan	Asia	Yamagata / Iwate	Tohoku	Regional	Delaware, Muscat Bailey A, Pinot Noir	Northern Honshū	Cool
India	Asia	Maharashtra	Nashik	GI	Sauvignon Blanc, Chenin, Shiraz, Cabernet	'Wine capital of India'	Tropical highland
India	Asia	Karnataka	Nandi Hills / Bangalore	GI	Shiraz, Cabernet, Sauvignon Blanc	High Deccan	Tropical highland
South Korea	Asia	Gyeonggi / Yeongdong	Yeongdong / Anseong	Regional	Campbell Early, MBA, Green Light	Nascent table and wine grapes	Continental, humid
Georgia	Europe	Kakheti	Napareuli	PDO	Saperavi, Rkatsiteli	Left-bank Kakheti	Continental
United States	North America	California	Napa Valley (other AVAs)	AVA	Cabernet Sauvignon, Chardonnay	Collective Napa floor and hills	Mediterranean
Chile	South America	Central Valley	Maipo Andes / Maipo Costa	Zona	Cabernet Sauvignon, Carmenère, Syrah	Upper Maipo vs coastal	Warm to cool
Argentina	South America	Salta	Cafayate / Calchaquíes	GI	Torrontés, Malbec, Cabernet, Tannat	Some of the world's highest vines	Extreme altitude desert
"""

COUNTRY_META = {
    "France": dict(continent="Europe", system="AOC / AOP / IGP", famous="Cabernet, Merlot, Pinot Noir, Chardonnay, Syrah, Chenin, Riesling", climate="Maritime to Mediterranean and alpine", note="World's most influential fine-wine culture. 288 appellations in this workbook from EU PDO + vineyards.com."),
    "Italy": dict(continent="Europe", system="DOCG / DOC / IGT", famous="Sangiovese, Nebbiolo, Barbera, Glera, Nero d'Avola, Aglianico", climate="Alpine north to hot Mediterranean south and islands", note="Most grape varieties of any country. 20 regions and 390 PDO appellations in this workbook."),
    "Spain": dict(continent="Europe", system="DOCa / DO / VCIG / VP / IGP", famous="Tempranillo, Garnacha, Albariño, Palomino, Monastrell, Verdejo", climate="Atlantic Galicia to hot continental Meseta and Mediterranean", note="Largest vineyard area on earth. Rioja and Priorat are DOCa; Sherry is unique."),
    "Portugal": dict(continent="Europe", system="DOC / Vinho Regional", famous="Touriga Nacional, Touriga Franca, Alvarinho, Baga, Castelão", climate="Atlantic green north to hot Alentejo; volcanic islands", note="Port, Madeira and a wave of dry table wines from indigenous grapes."),
    "Germany": dict(continent="Europe", system="Prädikat / VDP Grosse Lage / Qualitätswein", famous="Riesling, Spätburgunder, Silvaner", climate="Cool continental; steep river valleys", note="Riesling benchmark. 13 Anbaugebiete."),
    "Austria": dict(continent="Europe", system="DAC / Qualitätswein", famous="Grüner Veltliner, Riesling, Blaufränkisch, Zweigelt", climate="Cool continental and Pannonian east", note="Dry whites of precision; sweet wines from Neusiedlersee."),
    "United States": dict(continent="North America", system="AVA", famous="Cabernet Sauvignon, Pinot Noir, Chardonnay, Zinfandel, Riesling", climate="Huge range: foggy coasts to desert interiors", note="Napa, Sonoma, Willamette and Columbia Valley lead; 4+ in CruAtlas."),
    "Australia": dict(continent="Oceania", system="GI (Geographical Indication)", famous="Shiraz, Cabernet, Chardonnay, Riesling, Semillon, Grenache", climate="Warm Mediterranean to cool Tasmania", note="Old-vine Shiraz, Margaret River elegance, Hunter Semillon."),
    "New Zealand": dict(continent="Oceania", system="GI / Regional", famous="Sauvignon Blanc, Pinot Noir, Chardonnay, Syrah", climate="Cool maritime; Central Otago continental", note="Marlborough defined modern Sauvignon Blanc."),
    "Argentina": dict(continent="South America", system="DOC / IG / IP", famous="Malbec, Torrontés, Cabernet Franc, Bonarda", climate="High-altitude desert, Andean snowmelt irrigation", note="Highest commercial vineyards in Salta; Uco Valley is the quality engine."),
    "Chile": dict(continent="South America", system="DO (Denominación de Origen)", famous="Cabernet Sauvignon, Carmenère, Sauvignon Blanc, Syrah, País", climate="Mediterranean, squeezed by Pacific and Andes; north is desert", note="Phylloxera-free; old País and Carignan in the south."),
    "South Africa": dict(continent="Africa", system="WO (Wine of Origin)", famous="Chenin Blanc, Cabernet, Syrah, Pinotage, Sauvignon Blanc", climate="Mediterranean Cape, cool Hemel-en-Aarde to warm Klein Karoo", note="Oldest New World industry; Swartland old vines."),
    "United Kingdom": dict(continent="Europe", system="PDO / PGI", famous="Chardonnay, Pinot Noir, Pinot Meunier, Bacchus", climate="Cool maritime, warming", note="Sparkling on chalk Downs now rivals Champagne in blind tastings."),
    "Greece": dict(continent="Europe", system="PDO / PGI", famous="Assyrtiko, Xinomavro, Agiorgitiko, Moschofilero", climate="Mediterranean islands to cool Macedonia", note="Ancient wine culture, strong indigenous revival."),
    "Hungary": dict(continent="Europe", system="PDO / PGI / Aszú", famous="Furmint, Hárslevelű, Kékfrankos, Kadarka", climate="Continental; Tokaj botrytis", note="Tokaji Aszú is one of the world's great sweet wines."),
    "Switzerland": dict(continent="Europe", system="AOC / DOC", famous="Chasselas, Pinot Noir, Merlot, Petite Arvine", climate="Alpine valleys, lake-moderated", note="Tiny production, mostly drunk at home; Lavaux is UNESCO."),
    "Georgia": dict(continent="Europe", system="PDO / qvevri method", famous="Saperavi, Rkatsiteli, Mtsvane, Kisi, Chinuri", climate="Continental Kakheti, humid west", note="8,000 years of wine; UNESCO qvevri tradition."),
    "Slovenia": dict(continent="Europe", system="PTP / PGI", famous="Rebula, Rebula, Malvazija, Refošk, Furmint", climate="Sub-Mediterranean west, continental east", note="Brda continues Collio; orange-wine pioneer."),
    "Croatia": dict(continent="Europe", system="PDO / PGI", famous="Plavac Mali, Pošip, Malvazija, Graševina, Teran", climate="Mediterranean Dalmatia to continental Slavonia", note="Zinfandel's ancestral home (Crljenak/Tribidrag)."),
    "Romania": dict(continent="Europe", system="DOC / IG", famous="Fetească Neagră, Fetească Albă, Tămâioasă, Grasă de Cotnari", climate="Continental; Black Sea Dobrogea", note="Large vineyard area, improving quality in Dealu Mare."),
    "Bulgaria": dict(continent="Europe", system="PGI / PGI regions", famous="Mavrud, Melnik, Rubin, Cabernet, Merlot", climate="Continental to warm Thracian", note="Historic volume exporter, indigenous revival."),
    "Moldova": dict(continent="Europe", system="IG / PDO", famous="Rara Neagră, Fetească, Cabernet, Merlot", climate="Continental", note="One of the highest vineyard-per-capita figures; Mileștii Mici cellars."),
    "Czech Republic": dict(continent="Europe", system="VOC / Jakostní", famous="Riesling, Grüner Veltliner, Pálava, St. Laurent", climate="Cool Moravia", note="Moravia is ~96% of Czech wine."),
    "Slovakia": dict(continent="Europe", system="TOC / VOC", famous="Grüner Veltliner, Riesling, Dunaj, Frankovka", climate="Cool continental", note="Shares Tokaj with Hungary."),
    "Luxembourg": dict(continent="Europe", system="AOP Moselle", famous="Riesling, Pinot Gris, Auxerrois, Crémant", climate="Cool Moselle", note="Almost all vineyards on the Moselle terraces."),
    "Belgium": dict(continent="Europe", system="AOP / IGP", famous="Chardonnay, Pinot Noir", climate="Cool maritime", note="Tiny, sparkling-led."),
    "Netherlands": dict(continent="Europe", system="PDO", famous="Auxerrois, Johanniter, Rondo, Souvignier Gris", climate="Cool maritime", note="Hybrids and Pinot family on river terraces."),
    "North Macedonia": dict(continent="Europe", system="PGI", famous="Vranec, Smederevka, Tamjanika", climate="Hot continental Tikveš", note="Vranec is the signature red."),
    "Serbia": dict(continent="Europe", system="PGI", famous="Prokupac, Tamjanika, Riesling", climate="Continental", note="Fruška Gora and Šumadija leading a revival."),
    "Montenegro": dict(continent="Europe", system="PGI", famous="Vranac, Krstač", climate="Mediterranean-continental", note="Lake Skadar vineyards."),
    "Bosnia and Herzegovina": dict(continent="Europe", system="PGI", famous="Žilavka, Blatina", climate="Mediterranean inland", note="Herzegovina around Mostar."),
    "Albania": dict(continent="Europe", system="Regional", famous="Shesh i Zi, Shesh i Bardhë, Kallmet", climate="Mediterranean", note="Ancient Illyrian wine culture, small exports."),
    "Kosovo": dict(continent="Europe", system="Regional", famous="Vranac, Prokupac", climate="Continental", note="Rahovec is the historic vineyard zone."),
    "Ukraine": dict(continent="Europe", system="PGI", famous="Odesa Black, Rkatsiteli, Cabernet, sparkling", climate="Continental, Black Sea", note="Odesa and Transcarpathia; production disrupted by war."),
    "Armenia": dict(continent="Europe", system="GI", famous="Areni, Voskehat", climate="High continental", note="Areni-1 cave: oldest known winery (~4000 BCE)."),
    "Turkey": dict(continent="Europe", system="Regional", famous="Öküzgözü, Boğazkere, Kalecik Karası, Narince, Emir", climate="Aegean, Anatolian, Thracian", note="Anatolia is a grape-domestication cradle."),
    "Cyprus": dict(continent="Europe", system="PDO", famous="Xynisteri, Mavro, Maratheftiko", climate="Hot Mediterranean, Troodos altitude", note="Commandaria may be the oldest named wine still made."),
    "Lebanon": dict(continent="Asia", system="GI", famous="Cinsault, Cabernet, Syrah, Obaideh, Merwah", climate="High Bekaa Valley", note="Phoenician roots; Musar made the region famous."),
    "Israel": dict(continent="Asia", system="GI", famous="Cabernet, Merlot, Syrah, Carignan, Marselan", climate="Mediterranean to high desert", note="Biblical wine revival; Galilee and Golan lead."),
    "Canada": dict(continent="North America", system="VQA / GI", famous="Riesling, Vidal (Icewine), Chardonnay, Pinot Noir, Cabernet Franc", climate="Cool continental; Okanagan desert", note="Icewine icon; table wines now the story in Niagara and Okanagan."),
    "Mexico": dict(continent="North America", system="GI", famous="Cabernet, Tempranillo, Chenin, Grenache, Nebbiolo", climate="Warm, dry, Pacific nights in Baja", note="Valle de Guadalupe is the quality engine; Casa Madero 1597."),
    "Uruguay": dict(continent="South America", system="GI", famous="Tannat, Albariño, Merlot", climate="Humid Atlantic", note="Basque-founded Tannat culture."),
    "Brazil": dict(continent="South America", system="DO / GI", famous="Merlot, Chardonnay, sparkling, Tannat, Moscato", climate="Highland humid south; tropical São Francisco", note="Serra Gaúcha sparkling; two harvests in the north."),
    "Peru": dict(continent="South America", system="GI", famous="Quebranta, Italia (Pisco), emerging still wines", climate="Coastal desert", note="Pisco is the historic product; still wine is nascent."),
    "Bolivia": dict(continent="South America", system="GI", famous="Tannat, Malbec, Muscat of Alexandria", climate="High Andean valleys", note="Among the world's highest commercial vineyards."),
    "Morocco": dict(continent="Africa", system="AOG / AOC", famous="Syrah, Cabernet, Grenache, Vin Gris", climate="Mediterranean, Atlas altitude", note="Meknès and Atlas foothills."),
    "Tunisia": dict(continent="Africa", system="AOC", famous="Muscat d'Alexandrie, Carignan, Grenache", climate="Hot Mediterranean", note="Cap Bon and Mornag."),
    "Algeria": dict(continent="Africa", system="Regional", famous="Carignan, Cinsault, Alicante", climate="Mediterranean", note="Once a giant of French colonial production."),
    "China": dict(continent="Asia", system="GI / regional", famous="Cabernet Sauvignon, Marselan, Merlot, Chardonnay", climate="Continental desert (Ningxia) to humid coast and Himalayan Yunnan", note="Ningxia Helan Mountain is the quality leader."),
    "Japan": dict(continent="Asia", system="GI", famous="Koshu, Muscat Bailey A, Pinot Noir, Kerner", climate="Humid temperate; Hokkaidō cool", note="Koshu is the signature white; rain is the challenge."),
    "India": dict(continent="Asia", system="GI", famous="Sauvignon Blanc, Chenin, Shiraz, Cabernet", climate="Tropical highland (Nashik, Nandi Hills)", note="Monsoon-timed viticulture."),
    "South Korea": dict(continent="Asia", system="Regional", famous="Campbell Early, MBA", climate="Humid continental", note="Mostly table grapes; tiny wine industry."),
}


def parse_world_tsv():
    rows = []
    for line in WORLD_TSV.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split("\t")
        if len(parts) < 8:
            raise ValueError(f"Bad TSV line ({len(parts)} cols): {line[:80]}")
        country, continent, macro, region, klass, grapes, style, climate = parts[:8]
        rows.append(
            {
                "country": country,
                "continent": continent,
                "macro": macro,
                "region": region,
                "classification": klass,
                "grapes": grapes,
                "style": style,
                "climate": climate,
                "source": "World reference catalog",
                "in_atlas": "No",
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def style_header(ws, row, cols, bg=WINE, fg=CREAM):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(bg)
        cell.font = font(11, bold=True, color=fg)
        cell.alignment = align("center", "center", True)
        cell.border = thin


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, name, rows, cols):
    ref = f"A1:{get_column_letter(cols)}{rows}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    # Unique names required
    existing = {t.name for t in ws.parent.worksheets for t in getattr(ws, "tables", {}).values()} if False else set()
    ws.add_table(table)


def paint_data_rows(ws, start, end, cols, alt=ROW_ALT):
    for r in range(start, end + 1):
        bg = alt if r % 2 == 0 else WHITE
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            if cell.fill.fgColor is None or cell.fill.fgColor.rgb in ("00000000", "0"):
                cell.fill = fill(bg)
            cell.font = font(10)
            cell.alignment = align("left", "center", True)
            cell.border = thin


def sheet_about(wb, stats):
    ws = wb.active
    ws.title = "About"
    ws.sheet_properties.tabColor = WINE
    ws.hide_gridlines = True
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(0.5, 0.5, 0.6, 0.5)
    ws.merge_cells("B2:G2")
    ws.merge_cells("B3:G3")
    ws.merge_cells("B5:G5")
    ws["B2"] = "CRUATLAS"
    ws["B2"].font = Font(name="Georgia", size=28, bold=True, color=WINE)
    ws["B3"] = "World Wine Regions — reference workbook"
    ws["B3"].font = Font(name="Georgia", size=16, italic=True, color=SLATE)
    ws["B5"] = (
        "A working catalogue of wine-producing countries and their regions. "
        "France and Italy include the full PDO / appellation lists used by the CruAtlas map. "
        "Other countries list the principal GI / DO / AVA / DAC regions a sommelier or importer would expect. "
        "Use the filters on each sheet. This is a reference companion to the atlas — not a legal register."
    )
    ws["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B5"].font = font(12, color=INK)
    ws.row_dimensions[5].height = 72

    headers = ["Sheet", "What is in it", "Rows"]
    for i, h in enumerate(headers, 2):
        cell = ws.cell(7, i, h)
        cell.fill = fill(WINE)
        cell.font = font(11, bold=True, color=CREAM)
        cell.alignment = align("center")

    about_rows = [
        ("Countries", "Every producing country in this book, with classification system and signature grapes", stats["countries"]),
        ("All Regions", "Master list — macros + appellations worldwide, with grapes, style and climate", stats["all_regions"]),
        ("CruAtlas Detail", "The regions currently written up in the live atlas (climate, soil, water, vines)", stats["atlas"]),
        ("France Appellations", "Named French AOCs / PDOs from the map layer (vineyards.com + EU PDO)", stats["france"]),
        ("Italy Appellations", "Italian DOC / DOCG / PDO polygons from the map layer", stats["italy"]),
        ("Grapes", "Signature grape × region index (from atlas write-ups and appellation lists)", stats["grapes"]),
        ("Classification Guide", "What AOC, DOCG, AVA, GI, WO, DAC… actually mean", stats["classes"]),
    ]
    for i, (a, b, c) in enumerate(about_rows, 8):
        ws.cell(i, 2, a).font = font(11, bold=True, color=WINE)
        ws.cell(i, 3, b).font = font(11)
        ws.cell(i, 3).alignment = align("left", "center", True)
        ws.cell(i, 4, c).font = font(11, bold=True)
        ws.cell(i, 4).alignment = align("center")
        for col in range(2, 5):
            ws.cell(i, col).fill = fill(CREAM if i % 2 == 0 else WHITE)
            ws.cell(i, col).border = thin
        ws.row_dimensions[i].height = 28

    ws.merge_cells("B16:G16")
    ws["B16"] = "How to read it"
    ws["B16"].font = Font(name="Georgia", size=14, bold=True, color=WINE)

    notes = [
        "Country — sovereign wine-producing country (England is listed under United Kingdom).",
        "Macro region — the large zone you would point to on a wall map (Bordeaux, Tuscany, Mendoza, Napa Valley).",
        "Region / appellation — the named GI you would see on a label (Pauillac, Barolo, Uco Valley, Sta. Rita Hills).",
        "Classification — the legal or customary scheme (AOC, DOCG, DO, AVA, GI, WO, DAC, PDO…). See the guide sheet.",
        "In CruAtlas — Yes means the live app already has a write-up and (for France/Italy/Spain) a map polygon.",
        "Sources — CruAtlas write-ups; France & Italy EU PDO / vineyards.com layers; world catalogue compiled as a sommelier desk reference.",
        "Not exhaustive — village premiers crus, tiny pagos, and every US AVA are not all listed. The goal is complete countries and the regions that matter.",
    ]
    for i, n in enumerate(notes, 17):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        ws.cell(i, 2, "•  " + n).font = font(11)
        ws.cell(i, 2).alignment = align("left", "center", True)
        ws.row_dimensions[i].height = 22

    ws.merge_cells("B25:G25")
    ws["B25"] = "CruAtlas  ·  generated for the two-week atlas build  ·  filter any header to explore"
    ws["B25"].font = font(10, italic=True, color=MUTED)

    autosize(ws, [3, 26, 88, 14, 14, 14, 14])
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "B8"
    ws.print_title_rows = "1:7"
    ws.oddHeader.left.text = "CruAtlas · World Wine Regions"
    return ws


def write_table_sheet(wb, title, headers, rows, widths, tab_color=WINE, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(0.4, 0.4, 0.5, 0.5)
    ws.oddHeader.left.text = f"CruAtlas · {title}"
    ws.oddFooter.right.text = "Page &P of &N"
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.fill = fill(WINE)
        cell.font = font(11, bold=True, color=CREAM)
        cell.alignment = align("center", "center", True)
    ws.row_dimensions[1].height = 28
    for r_i, row in enumerate(rows, 2):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(r_i, c_i, val)
            cell.font = font(10)
            cell.alignment = align("left", "center", True)
            cell.border = thin
            if r_i % 2 == 0:
                cell.fill = fill(ROW_ALT)
        ws.row_dimensions[r_i].height = 32 if title in ("CruAtlas Detail", "About") else 22
    autosize(ws, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows)+1)}"
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"
    safe = re.sub(r"[^A-Za-z0-9]", "", title) or "Table"
    if rows:
        tbl = Table(displayName=f"T_{safe}", ref=f"A1:{get_column_letter(len(headers))}{1 + len(rows)}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)
    ws.print_title_rows = "1:1"
    return ws


def main():
    countries_meta, atlas = parse_atlas()
    france = load_geojson(ROOT / "public/data/france-appellations.geojson")
    italy = load_geojson(ROOT / "public/data/italy-subregions.geojson")
    italy_macro = load_geojson(ROOT / "public/data/italy-wine-regions.geojson")
    world = parse_world_tsv()

    atlas_keys = {(r["country"], r["name"].lower()) for r in atlas}
    atlas_names_fr = {r["name"].lower() for r in atlas if r["country"] == "France"}
    atlas_names_it = {r["name"].lower() for r in atlas if r["country"] == "Italy"}

    # Master region rows
    master = []

    def add_master(**kwargs):
        master.append(kwargs)

    # Atlas macros (all 76)
    MACRO_PARENT = {
        "France": {
            "Médoc / Haut-Médoc": "Bordeaux",
            "Pauillac": "Bordeaux",
            "Margaux": "Bordeaux",
            "Saint-Julien": "Bordeaux",
            "Saint-Estèphe": "Bordeaux",
            "Pessac-Léognan": "Bordeaux",
            "Sauternes & Barsac": "Bordeaux",
            "Saint-Émilion": "Bordeaux",
            "Pomerol": "Bordeaux",
            "Fronsac & Canon-Fronsac": "Bordeaux",
            "Graves": "Bordeaux",
            "Entre-Deux-Mers": "Bordeaux",
            "Bourg & Blaye": "Bordeaux",
            "Chablis": "Burgundy",
            "Irancy & Saint-Bris": "Burgundy",
            "Côte de Nuits": "Burgundy",
            "Côte de Beaune": "Burgundy",
            "Côte Chalonnaise": "Burgundy",
            "Mâconnais": "Burgundy",
            "Northern Rhône": "Rhône",
            "Châteauneuf-du-Pape & Southern Rhône": "Rhône",
            "Sancerre & Pouilly-Fumé": "Loire",
            "Vouvray & Central Loire": "Loire",
            "Muscadet (Nantais)": "Loire",
            "Anjou-Saumur": "Loire",
            "Chinon & Bourgueil": "Loire",
            "Bandol": "Provence",
            "Cassis": "Provence",
        }
    }

    continent_of = {c: m["continent"] for c, m in COUNTRY_META.items()}
    continent_of.update({"United States": "North America", "South Africa": "Africa"})

    for r in atlas:
        macro = MACRO_PARENT.get(r["country"], {}).get(r["name"], r["name"] if r["country"] != "France" else r["name"])
        if r["country"] == "France" and r["name"] not in MACRO_PARENT["France"]:
            macro = r["name"]
        add_master(
            country=r["country"],
            continent=continent_of.get(r["country"], "Europe"),
            macro=MACRO_PARENT.get(r["country"], {}).get(r["name"], r["name"]),
            region=r["name"],
            classification="Atlas write-up",
            grapes=", ".join(r["vines"]),
            style=r["desc"],
            climate=r["climate"],
            source="CruAtlas",
            in_atlas="Yes",
            level="Macro / featured",
        )

    for p in france:
        add_master(
            country="France",
            continent="Europe",
            macro=p["region"] or "France",
            region=p["name"],
            classification="AOC / AOP",
            grapes=", ".join(p["grapes"][:8]),
            style=p["desc"],
            climate="",
            source="France PDO layer",
            in_atlas="Yes" if p["name"].lower() in atlas_names_fr or p["region"].lower() in {x.lower() for x in atlas_names_fr} else "Map only",
            level="Appellation",
        )

    for p in italy_macro:
        add_master(
            country="Italy",
            continent="Europe",
            macro=p["name"],
            region=p["name"],
            classification="Regione",
            grapes=", ".join(p["grapes"][:8]),
            style=p["desc"],
            climate="",
            source="Italy regions layer",
            in_atlas="Yes" if p["name"].lower() in atlas_names_it else "Map only",
            level="Macro region",
        )

    for p in italy:
        add_master(
            country="Italy",
            continent="Europe",
            macro=p["region"] or "Italy",
            region=p["name"],
            classification="DOC / DOCG / PDO",
            grapes=", ".join(p["grapes"][:8]),
            style=p["desc"],
            climate="",
            source="Italy PDO layer",
            in_atlas="Yes" if p["name"].lower() in atlas_names_it else "Map only",
            level="Appellation",
        )

    for w in world:
        add_master(
            country=w["country"],
            continent=w["continent"],
            macro=w["macro"],
            region=w["region"],
            classification=w["classification"],
            grapes=w["grapes"],
            style=w["style"],
            climate=w["climate"],
            source=w["source"],
            in_atlas="Yes" if (w["country"], w["region"].lower()) in atlas_keys else "No",
            level="Region / GI",
        )

    # Deduplicate by country+region, prefer CruAtlas / PDO over world
    priority = {"CruAtlas": 0, "France PDO layer": 1, "Italy PDO layer": 1, "Italy regions layer": 2, "World reference catalog": 3}
    master.sort(key=lambda x: (x["country"], x["macro"], x["region"], priority.get(x["source"], 9)))
    seen = set()
    unique = []
    for row in master:
        key = (row["country"].lower(), row["region"].lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    master = unique
    master.sort(key=lambda x: (x["continent"], x["country"], x["macro"], x["region"]))

    # Countries sheet
    region_counts = Counter(r["country"] for r in master)
    atlas_counts = Counter(r["country"] for r in atlas)
    fr_count = sum(1 for r in master if r["country"] == "France")
    it_count = sum(1 for r in master if r["country"] == "Italy")

    country_rows_data = []
    # Union of COUNTRY_META and anything in master
    all_countries = sorted(set(COUNTRY_META) | {r["country"] for r in master}, key=lambda c: (continent_of.get(c, "ZZ"), c))
    flags = {n: f for _, i, n, f in countries_meta}
    flags.update({
        "France": "🇫🇷", "Italy": "🇮🇹", "Spain": "🇪🇸", "Portugal": "🇵🇹", "Germany": "🇩🇪",
        "Austria": "🇦🇹", "United States": "🇺🇸", "Australia": "🇦🇺", "New Zealand": "🇳🇿",
        "Argentina": "🇦🇷", "Chile": "🇨🇱", "South Africa": "🇿🇦", "United Kingdom": "🇬🇧",
        "Greece": "🇬🇷", "Hungary": "🇭🇺", "Switzerland": "🇨🇭", "Georgia": "🇬🇪",
        "Slovenia": "🇸🇮", "Croatia": "🇭🇷", "Romania": "🇷🇴", "Bulgaria": "🇧🇬",
        "Moldova": "🇲🇩", "Czech Republic": "🇨🇿", "Slovakia": "🇸🇰", "Luxembourg": "🇱🇺",
        "Belgium": "🇧🇪", "Netherlands": "🇳🇱", "North Macedonia": "🇲🇰", "Serbia": "🇷🇸",
        "Montenegro": "🇲🇪", "Bosnia and Herzegovina": "🇧🇦", "Albania": "🇦🇱", "Kosovo": "🇽🇰",
        "Ukraine": "🇺🇦", "Armenia": "🇦🇲", "Turkey": "🇹🇷", "Cyprus": "🇨🇾",
        "Lebanon": "🇱🇧", "Israel": "🇮🇱", "Canada": "🇨🇦", "Mexico": "🇲🇽",
        "Uruguay": "🇺🇾", "Brazil": "🇧🇷", "Peru": "🇵🇪", "Bolivia": "🇧🇴",
        "Morocco": "🇲🇦", "Tunisia": "🇹🇳", "Algeria": "🇩🇿", "China": "🇨🇳",
        "Japan": "🇯🇵", "India": "🇮🇳", "South Korea": "🇰🇷",
    })

    for c in all_countries:
        meta = COUNTRY_META.get(c, dict(continent=continent_of.get(c, ""), system="", famous="", climate="", note=""))
        country_rows_data.append([
            flags.get(c, ""),
            c,
            meta.get("continent", ""),
            meta.get("system", ""),
            region_counts.get(c, 0),
            atlas_counts.get(c, 0),
            meta.get("famous", ""),
            meta.get("climate", ""),
            meta.get("note", ""),
        ])

    # Grapes index
    grape_rows = []
    for r in atlas:
        for g in r["vines"]:
            grape_rows.append([g, r["country"], r["name"], "CruAtlas write-up", r["desc"][:180]])
    for p in france:
        for g in p["grapes"][:6]:
            grape_rows.append([g, "France", p["name"], p["region"], "AOC"])
    for p in italy:
        for g in p["grapes"][:6]:
            grape_rows.append([g, "Italy", p["name"], p["region"], "DOC/DOCG"])
    for w in world:
        for g in [x.strip() for x in w["grapes"].split(",") if x.strip()]:
            # skip overly long lists dumped as one grape
            if len(g) > 48:
                continue
            grape_rows.append([g, w["country"], w["region"], w["macro"], w["classification"]])
    # unique grape-country-region
    gseen = set()
    grape_unique = []
    for row in grape_rows:
        key = (row[0].lower(), row[1].lower(), row[2].lower())
        if key in gseen:
            continue
        gseen.add(key)
        grape_unique.append(row)
    grape_unique.sort(key=lambda x: (x[0].lower(), x[1], x[2]))

    class_rows = [
        ["AOC / AOP", "France", "Appellation d'Origine Contrôlée / Protégée — the classic French GI. AOP is the EU name for the same idea."],
        ["IGP / Vin de Pays", "France", "Indication Géographique Protégée — broader than AOC (e.g. Pays d'Oc)."],
        ["DOCG", "Italy", "Denominazione di Origine Controllata e Garantita — top Italian tier, with tasting panel (Barolo, Brunello, Amarone…)."],
        ["DOC", "Italy / Portugal", "Denominazione / Denominação de Origine Controllata — the standard protected tier."],
        ["IGT / IGP", "Italy", "Indicazione Geografica Tipica — the Super-Tuscan home; looser grape and style rules."],
        ["DOCa / DOQ", "Spain", "Denominación de Origen Calificada — currently Rioja and Priorat (Catalan: DOQ)."],
        ["DO", "Spain / Chile", "Denominación de Origen — standard Spanish (and Chilean) GI."],
        ["VCIG / VP / VT", "Spain", "Vino de Calidad, Vino de Pago (single estate), Vino de la Tierra (IGP)."],
        ["Sherry / Jerez", "Spain", "Fortified PDO with biological (Fino/Manzanilla) and oxidative (Oloroso) ageing in solera."],
        ["DOC / VR", "Portugal", "DOC is the top Portuguese GI; Vinho Regional is the IGP equivalent."],
        ["Port / Madeira", "Portugal", "Fortified PDOs with their own institutes (IVDP, IVBAM)."],
        ["Anbaugebiet / Prädikat", "Germany", "13 regions; Prädikat (Kabinett → Trockenbeerenauslese) is ripeness, not quality by itself. VDP Grosse Lage is the leading dry/sweet site classification."],
        ["DAC", "Austria", "Districtus Austriae Controllatus — region + grape + style, similar to AOC."],
        ["PDO / PGI", "EU / UK", "Umbrella EU (and post-Brexit UK) legal terms. AOC/DOCG/DO are all PDOs."],
        ["AVA", "United States", "American Viticultural Area — a geographic claim only; no grape or yield rules."],
        ["VQA", "Canada", "Vintners Quality Alliance — Ontario and BC appellation + tasting panel."],
        ["GI", "Australia / NZ / others", "Geographical Indication — the New World AOC analogue, usually geographic only."],
        ["WO", "South Africa", "Wine of Origin — estate / ward / district / region, with a certification seal."],
        ["Qvevri / traditional method", "Georgia", "UNESCO-listed clay-vessel fermentation; PDOs sit on top of grape and place rules."],
        ["Tokaji Aszú / puttonyos", "Hungary", "Botrytis dessert classification (now mostly 5–6 puttonyos and Eszencia)."],
        ["Commandaria", "Cyprus", "Sun-dried sweet PDO, among the oldest named wines still in production."],
        ["Icewine / Eiswein", "Canada / Germany / Austria", "Grapes frozen on the vine; Canada VQA is the volume leader."],
    ]

    stats = {
        "countries": len(all_countries),
        "all_regions": len(master),
        "atlas": len(atlas),
        "france": len(france),
        "italy": len(italy),
        "grapes": len(grape_unique),
        "classes": len(class_rows),
    }

    wb = Workbook()
    sheet_about(wb, stats)

    write_table_sheet(
        wb,
        "Countries",
        ["Flag", "Country", "Continent / zone", "Classification system", "Regions in book", "In CruAtlas", "Signature grapes", "Climate snapshot", "Notes"],
        country_rows_data,
        [8, 26, 20, 36, 16, 14, 48, 42, 62],
        tab_color=SAGE,
    )

    write_table_sheet(
        wb,
        "All Regions",
        ["Country", "Continent", "Macro region", "Region / appellation", "Level", "Classification", "Signature grapes", "Style / notes", "Climate", "In CruAtlas", "Source"],
        [
            [
                r["country"],
                r["continent"],
                r["macro"],
                r["region"],
                r["level"],
                r["classification"],
                r["grapes"],
                r["style"],
                r["climate"],
                r["in_atlas"],
                r["source"],
            ]
            for r in master
        ],
        [18, 16, 24, 36, 16, 22, 42, 56, 36, 14, 22],
        tab_color=WINE,
    )

    write_table_sheet(
        wb,
        "CruAtlas Detail",
        ["Country", "Region", "Lat", "Lng", "Grapes", "Climate", "Soil", "Water / irrigation", "Typical wines", "Description"],
        [
            [
                r["country"],
                r["name"],
                r["lat"],
                r["lng"],
                ", ".join(r["vines"]),
                r["climate"],
                r["soil"],
                r["water"],
                ", ".join(r["wines"]),
                r["desc"],
            ]
            for r in atlas
        ],
        [18, 32, 10, 10, 42, 48, 48, 40, 42, 56],
        tab_color=GOLD,
    )
    # slightly taller rows for the detail sheet
    det = wb["CruAtlas Detail"]
    for r in range(2, 2 + len(atlas)):
        det.row_dimensions[r].height = 48

    write_table_sheet(
        wb,
        "France Appellations",
        ["Appellation", "Macro region", "Grapes", "Description", "Id"],
        [[p["name"], p["region"], ", ".join(p["grapes"][:10]), p["desc"], p["id"]] for p in sorted(france, key=lambda x: (x["region"], x["name"]))],
        [36, 24, 48, 64, 28],
        tab_color="7A1F2B",
    )

    write_table_sheet(
        wb,
        "Italy Appellations",
        ["Appellation", "Region (regione)", "Grapes", "Description", "Level", "Id"],
        [[p["name"], p["region"], ", ".join(p["grapes"][:10]), p["desc"], p["level"], p["id"]] for p in sorted(italy, key=lambda x: (x["region"], x["name"]))],
        [36, 24, 40, 56, 16, 28],
        tab_color="2F4A2C",
    )

    write_table_sheet(
        wb,
        "Grapes",
        ["Grape", "Country", "Region / appellation", "Macro / zone", "Context"],
        grape_unique,
        [32, 22, 36, 28, 48],
        tab_color=SAGE,
    )

    write_table_sheet(
        wb,
        "Classification Guide",
        ["Term", "Where", "What it means"],
        class_rows,
        [28, 28, 100],
        tab_color=SLATE,
    )
    cg = wb["Classification Guide"]
    for r in range(2, 2 + len(class_rows)):
        cg.row_dimensions[r].height = 36

    # Highlight Yes in In CruAtlas column of All Regions
    ar = wb["All Regions"]
    yes_fill = PatternFill("solid", fgColor="E4F0D8")
    no_fill = PatternFill("solid", fgColor="F6EDE4")
    map_fill = PatternFill("solid", fgColor="E8EEF6")
    for r in range(2, 2 + len(master)):
        cell = ar.cell(r, 10)
        if cell.value == "Yes":
            cell.fill = yes_fill
            cell.font = font(10, bold=True, color="2F4A2C")
        elif cell.value == "Map only":
            cell.fill = map_fill
            cell.font = font(10, color="1F3A5F")
        else:
            cell.fill = no_fill

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  countries {stats['countries']}")
    print(f"  all regions {stats['all_regions']}")
    print(f"  atlas {stats['atlas']}")
    print(f"  france {stats['france']}")
    print(f"  italy {stats['italy']}")
    print(f"  grapes {stats['grapes']}")
    print(f"  size {OUT.stat().st_size}")


if __name__ == "__main__":
    main()
